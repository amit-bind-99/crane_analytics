from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, session, flash
from flask_bcrypt import Bcrypt
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import r2_score
from werkzeug.utils import secure_filename
from functools import wraps
import sqlite3
import warnings
import os
import re
import io
import shutil
from datetime import datetime, timedelta
import json

warnings.filterwarnings('ignore')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'yz-bay-crane-analytics-secret-key-change-in-production')
bcrypt = Bcrypt(app)

DATA_DIR = 'data'
DATA_PATH = os.path.join(DATA_DIR, 'LT Wheel replacement data.xlsx')
SAMPLE_FILENAME = 'LT Wheel replacement data.xlsx'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
DB_PATH = os.path.join(DATA_DIR, 'crane_analytics.db')

# ============== CONFIGURATION ==============

REQUIRED_WHEEL_COLUMNS = {'Date', 'Crane', 'Equipment', 'Remarks'}
OPTIONAL_WHEEL_COLUMNS = {'S.no.', 'Job Description', 'Position'}
VALID_POSITIONS = {'SW', 'SE', 'NE', 'NW'}
VALID_CRANE_KEYWORDS = ['west', 'east']

# ============== DATABASE ==============

def get_db_connection():
    """Get a SQLite database connection with row factory."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize the database schema."""
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS excel_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            version_number INTEGER NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            total_wheel_records INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
        )
    ''')

    # Create default admin user if none exists
    cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
    if cursor.fetchone()[0] == 0:
        admin_hash = bcrypt.generate_password_hash('admin123').decode('utf-8')
        cursor.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            ('admin', admin_hash, 'admin')
        )

    conn.commit()
    conn.close()

def ensure_db():
    """Ensure DB schema exists; useful when the DB file may have been recreated."""
    try:
        conn = get_db_connection()
        conn.execute("SELECT 1 FROM users LIMIT 1")
        conn.close()
    except sqlite3.OperationalError:
        init_db()

init_db()

# ============== AUTH HELPERS ==============

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': 'Admin access required'}), 403
            flash('Admin access required', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    """Get current user from session as dict, or None."""
    if 'user_id' not in session:
        return None
    conn = get_db_connection()
    user = conn.execute(
        "SELECT id, username, role, is_active FROM users WHERE id = ?",
        (session['user_id'],)
    ).fetchone()
    conn.close()
    if user is None or not user['is_active']:
        session.clear()
        return None
    return dict(user)

# ============== MOCK DATA GENERATORS ==============

def generate_mock_wheel_data():
    """Generate realistic wheel replacement mock data in the new Excel format."""
    np.random.seed(42)
    cranes = ['YZ bay west crane', 'YZ bay East crane']
    positions = ['SW', 'NE', 'NW', 'SE']
    wheel_types = ['drive wheel', 'idle wheel']
    remarks_list = [
        'Shutdown job',
        'Breakdown job',
        'Planned maintenance',
        'Inspection replacement'
    ]

    data = []
    start_date = datetime(2024, 1, 1)
    end_date = datetime(2025, 9, 30)

    for i in range(87):
        crane = np.random.choice(cranes, p=[0.58, 0.42])
        pos = np.random.choice(positions)
        wheel_type = np.random.choice(wheel_types)
        equipment = f"LT {pos} {wheel_type}"

        days_offset = int(np.random.beta(2, 1.2) * (end_date - start_date).days)
        date = start_date + timedelta(days=days_offset)

        remark = np.random.choice(remarks_list)
        job_desc = (
            f"{crane} LT {pos} {wheel_type} replacement done. "
            f"(Wheel collar {'reduced' if remark == 'Shutdown job' else 'broken'})"
        )

        data.append({
            'S.no.': i + 1,
            'Date': date,
            'Crane': crane,
            'Equipment': equipment,
            'Job Description': job_desc,
            'Remarks': remark
        })

    df = pd.DataFrame(data)
    df = df.sort_values('Date').reset_index(drop=True)
    df['S.no.'] = range(1, len(df) + 1)
    return df

def generate_mock_hardness_data():
    """Generate realistic rail hardness data"""
    np.random.seed(7)
    sections = ['21-22', '22-23', '23-24', '24-25', '25-26', '26-27',
                '27-28', '28-29', '29-30', '30-31', '31-32', 'End portion 32-33']

    north = [270, 270, 330, 466, 374, 412, 420, 380, 380, 360, 320, 270]
    north = [h + np.random.randint(-8, 9) for h in north]

    south = [270, 270, 330, 395, 390, 370, 380, 400, 355, 380, 360, 270]
    south = [h + np.random.randint(-7, 8) for h in south]

    return pd.DataFrame({
        'Section': sections,
        'North': north,
        'South': south
    })

def generate_mock_rail_replacement_data():
    """Generate rail replacement log in the new Excel format."""
    data = [
        {
            'S.no.': 1,
            'Date': '29.07.2025',
            'Crane': 'YZ bay LT rail',
            'Notification no.': '1600543389 / 1600700448',
            'Equipment': 'YZ bay LT Rail',
            'Job Description': 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 29 to 31)',
            'Remarks': 'Shutdown job'
        },
        {
            'S.no.': 2,
            'Date': '29.07.2025',
            'Crane': 'YZ bay LT rail joint',
            'Notification no.': '1600543390 / 1600700449',
            'Equipment': 'YZ bay LT rail joint Thermit welding',
            'Job Description': 'YZ bay LT rail joints (4 nos.) Thermit welding done. (Column no. 29 to 31)',
            'Remarks': 'Shutdown job'
        },
        {
            'S.no.': 3,
            'Date': '30.07.2025',
            'Crane': 'YZ bay LT rail',
            'Notification no.': '1600543391 / 1600700540',
            'Equipment': 'YZ bay LT Rail',
            'Job Description': 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 26 to 29)',
            'Remarks': 'Shutdown job'
        },
        {
            'S.no.': 4,
            'Date': '30.07.2025',
            'Crane': 'YZ bay LT rail joint',
            'Notification no.': '1600543392 / 1600700541',
            'Equipment': 'YZ bay LT rail joint Thermit welding',
            'Job Description': 'YZ bay LT rail joints (4 nos.) Thermit welding done. (Column no. 26 to 29)',
            'Remarks': 'Shutdown job'
        },
        {
            'S.no.': 5,
            'Date': '31.07.2025',
            'Crane': 'YZ bay LT rail',
            'Notification no.': '1600543394 / 1600700543',
            'Equipment': 'YZ bay LT Rail',
            'Job Description': 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 23 to 26)',
            'Remarks': 'Shutdown job'
        }
    ]
    return pd.DataFrame(data)

# ============== HELPERS ==============

def safe_float(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

def get_risk_class(hb):
    if hb is None:
        return 'Unknown'
    if hb > 400:
        return 'Critical'
    if hb > 350:
        return 'High'
    if hb > 300:
        return 'Medium'
    return 'Normal'

def normalize_crane(crane):
    """Normalize crane names to LT WEST / LT EAST."""
    if pd.isna(crane):
        return 'UNKNOWN'
    c = str(crane).lower()
    if 'west' in c:
        return 'LT WEST'
    if 'east' in c:
        return 'LT EAST'
    return str(crane).strip()

def extract_position(row):
    """Extract wheel position (SW/SE/NE/NW) from Equipment or Job Description."""
    for col in ['Equipment', 'Job Description']:
        if col in row and pd.notna(row[col]):
            text = str(row[col]).upper()
            for pos in VALID_POSITIONS:
                if re.search(r'\b' + pos + r'\b', text):
                    return pos
    return 'Other'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_version_storage_path(version_id):
    """Return storage path for a specific Excel version."""
    return os.path.join(DATA_DIR, f'version_{version_id}.xlsx')

# ============== RAIL REPLACEMENT PARSERS ==============

def extract_rail_qty(job_description):
    """Extract quantity like '6 nos.' from job description.

    Counts only actual rail PIECE replacements; thermit welding rail
    joints are excluded from the piece count (return 0).
    """
    if pd.isna(job_description):
        return 0
    text = str(job_description)
    if re.search(r'thermit|welding|joint', text, re.IGNORECASE):
        return 0
    match = re.search(r'(\d+)\s*no\.?', text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 0

def extract_rail_section(job_description):
    """Extract section range like 'Column no. 29 to 31' from job description."""
    if pd.isna(job_description):
        return 'Unknown'
    text = str(job_description)
    match = re.search(r'Column no\.\s*(\d+)\s*to\s*(\d+)', text, re.IGNORECASE)
    if match:
        start, end = match.group(1), match.group(2)
        return f"Column {start} to {end}"
    match = re.search(r'Column no\.\s*(\d+)', text, re.IGNORECASE)
    if match:
        return f"Column {match.group(1)}"
    return 'Unknown'

def extract_rail_reason(equipment, job_description):
    """Derive reason from equipment or job description."""
    text = ' '.join([str(equipment), str(job_description)]).lower()
    if 'thermit' in text or 'welding' in text:
        return 'Thermit welding'
    if 'replacement' in text:
        return 'Rail pieces replacement'
    if 'crack' in text:
        return 'Crack repair'
    return 'Maintenance'

def parse_rail_replacement_dataframe(df_raw):
    """
    Normalize rail replacement dataframe to internal format:
    Date, Section, Side, Qty_Pieces, Reason
    Supports both old format and new format.
    """
    df = df_raw.copy()

    # Strip leading/trailing whitespace from column names
    df.columns = [str(c).strip() for c in df.columns]

    # Old format detection
    old_cols = {'Date', 'Section', 'Side', 'Qty_Pieces', 'Reason'}
    if old_cols.issubset(set(df.columns)):
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)
        df['Qty_Pieces'] = pd.to_numeric(df['Qty_Pieces'], errors='coerce').fillna(0).astype(int)
        return df[['Date', 'Section', 'Side', 'Qty_Pieces', 'Reason']]

    # New format
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)

    df['Qty_Pieces'] = df.apply(lambda r: extract_rail_qty(r.get('Job Description')), axis=1)
    df['Section'] = df.apply(lambda r: extract_rail_section(r.get('Job Description')), axis=1)
    df['Reason'] = df.apply(lambda r: extract_rail_reason(r.get('Equipment'), r.get('Job Description')), axis=1)
    df['Side'] = 'Both'

    return df[['Date', 'Section', 'Side', 'Qty_Pieces', 'Reason']]

# ============== VALIDATION ==============

def validate_wheel_dataframe(df, source_name='uploaded file'):
    """Lenient validation: only fail if the sheet is completely empty."""
    errors = []
    warnings_list = []
    stats = {}

    if df is None or df.empty:
        errors.append(f"{source_name} contains no data rows.")
        return {'valid': False, 'errors': errors, 'warnings': warnings_list, 'stats': stats}

    df = df.dropna(how='all')
    stats['total_rows'] = len(df)
    columns = set(df.columns)

    missing_required = REQUIRED_WHEEL_COLUMNS - columns
    if missing_required:
        warnings_list.append(f"Missing columns (will use defaults): {', '.join(sorted(missing_required))}.")

    if 'Date' in columns:
        df['Date_parsed'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)
        invalid_dates = int(df['Date_parsed'].isna().sum())
        if invalid_dates > 0:
            warnings_list.append(f"{invalid_dates} row(s) have invalid or missing dates (will be skipped in charts).")
        valid_dates = df['Date_parsed'].dropna()
        if not valid_dates.empty:
            stats['date_range'] = {
                'min': valid_dates.min().strftime('%Y-%m-%d'),
                'max': valid_dates.max().strftime('%Y-%m-%d')
            }

    if 'Crane' in columns:
        unique_cranes = df['Crane'].dropna().astype(str).apply(normalize_crane).unique().tolist()
        stats['cranes'] = unique_cranes

    if 'Equipment' in columns or 'Job Description' in columns:
        df['Position_extracted'] = df.apply(extract_position, axis=1)
        stats['positions'] = df['Position_extracted'].value_counts().to_dict()

    return {'valid': True, 'errors': [], 'warnings': warnings_list, 'stats': stats}

def validate_rail_replacement_dataframe(df, source_name='Rail Replacement data'):
    """Lenient validation: accept whatever columns are present."""
    if df is None or df.empty:
        return {'valid': True, 'errors': [], 'warnings': [f"{source_name} sheet is empty."], 'stats': {}}
    df = df.dropna(how='all')
    return {'valid': True, 'errors': [], 'warnings': [], 'stats': {'total_rows': len(df)}}
def validate_hardness_dataframe(df, source_name='Rail Hardness data'):
    """Lenient validation: correct row indices, NaN-safe stats."""
    if df is None or df.empty:
        return {'valid': True, 'errors': [], 'warnings': [f"{source_name} sheet is empty."], 'stats': {}}

    stats = {}
    try:
        # Layout: row 0 = empty, row 1 = section labels, row 2 = north, row 3 = south
        sections = df.iloc[1, 2:14].astype(str).tolist()
        north = pd.to_numeric(df.iloc[2, 2:14], errors='coerce')
        south = pd.to_numeric(df.iloc[3, 2:14], errors='coerce')
        stats['sections'] = len(sections)
        north_mean = north.dropna().mean()
        south_mean = south.dropna().mean()
        stats['north_avg'] = round(float(north_mean), 1) if pd.notna(north_mean) else None
        stats['south_avg'] = round(float(south_mean), 1) if pd.notna(south_mean) else None
    except Exception:
        stats['sections'] = 0

    return {'valid': True, 'errors': [], 'warnings': [], 'stats': stats}

# ============== SAMPLE EXCEL GENERATOR ==============

def generate_sample_excel():
    """Generate a sample Excel workbook with data validation for all 3 sheets."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ---------- Sheet 1: LT wheel replacement data ----------
    ws1 = wb.active
    ws1.title = 'LT wheel replacement data'

    headers1 = ['S.no.', 'Date', 'Crane', 'Equipment', 'Job Description', 'Remarks']
    ws1.append(headers1)

    header_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sample_wheel_rows = [
        [1, '02.04.2024', 'YZ bay west crane', 'LT SW drive wheel',
         'YZ bay west crane LT SW drive wheel replacement done. (Wheel collar reduced)', 'Shutdown job'],
        [2, '02.04.2024', 'YZ bay west crane', 'LT SE drive wheel',
         'YZ bay west crane LT SE drive wheel replacement done. (Wheel collar reduced)', 'Shutdown job'],
        [3, '06.04.2024', 'YZ bay East crane', 'LT NE idle wheel',
         'YZ bay east crane LT NE idle wheel replacement done. (Wheel collar reduced)', 'Shutdown job'],
        [4, '29.07.2024', 'YZ bay west crane', 'LT NE drive wheel',
         'YZ bay west crane LT NE drive wheel replacement done. (Wheel collar reduced)', 'Shutdown job'],
        [5, '03.12.2024', 'YZ bay west crane', 'LT SW idle wheel',
         'YZ bay west crane LT SW idle wheel replacement done. (Wheel collar reduced)', 'Planned maintenance'],
        [6, '13.02.2025', 'YZ bay East crane', 'LT NW drive wheel',
         'YZ bay east crane LT NW drive wheel replacement done. (Wheel collar broken)', 'Breakdown job'],
    ]

    for row in sample_wheel_rows:
        ws1.append(row)

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 65
    ws1.column_dimensions['F'].width = 20

    crane_validation = DataValidation(
        type="list",
        formula1='"YZ bay west crane,YZ bay East crane"',
        allow_blank=False
    )
    crane_validation.error = 'Please select a valid crane from the list.'
    crane_validation.errorTitle = 'Invalid Crane'
    crane_validation.prompt = 'Select the crane where the replacement occurred.'
    crane_validation.promptTitle = 'Crane'
    ws1.add_data_validation(crane_validation)
    crane_validation.add('C2:C1000')

    equipment_options = [
        'LT SW drive wheel', 'LT SW idle wheel',
        'LT SE drive wheel', 'LT SE idle wheel',
        'LT NE drive wheel', 'LT NE idle wheel',
        'LT NW drive wheel', 'LT NW idle wheel'
    ]
    equipment_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(equipment_options) + '"',
        allow_blank=False
    )
    equipment_validation.error = 'Please select a valid equipment value.'
    equipment_validation.errorTitle = 'Invalid Equipment'
    equipment_validation.prompt = 'Select the wheel equipment replaced.'
    equipment_validation.promptTitle = 'Equipment'
    ws1.add_data_validation(equipment_validation)
    equipment_validation.add('D2:D1000')

    remarks_validation = DataValidation(
        type="list",
        formula1='"Shutdown job,Breakdown job,Planned maintenance,Inspection replacement"',
        allow_blank=False
    )
    remarks_validation.error = 'Please select a valid remark category.'
    remarks_validation.errorTitle = 'Invalid Remarks'
    ws1.add_data_validation(remarks_validation)
    remarks_validation.add('F2:F1000')

    ws1.freeze_panes = 'A2'

    # ---------- Sheet 2: Rail Hardness data ----------
    ws2 = wb.create_sheet('Rail Hardness data')
    ws2.append(['', 'Column no.', '21-22', '22-23', '23-24', '24-25', '25-26',
                '26-27', '27-28', '28-29', '29-30', '30-31', '31-32', 'End portion 32-33'])
    ws2.append(['', 'North Side Z bay Rail hardness (HB)',
                270, 270, 330, 466, 374, 412, 420, 380, 380, 360, 320, 270])
    ws2.append(['', 'South Side Y bay Rail hardness (HB)',
                270, 270, 330, 395, 390, 370, 380, 400, 355, 380, 360, 270])

    for row in ws2.iter_rows(min_row=1, max_row=3, min_col=2, max_col=14):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    # ---------- Sheet 3: Rail Replacement data ----------
    ws3 = wb.create_sheet('Rail Replacement data')
    headers3 = ['S.no.', 'Date', 'Crane', 'Notification no.', 'Equipment', 'Job Description', 'Remarks']
    ws3.append(headers3)

    for col_num, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sample_rail_rows = [
        [1, '29.07.2025', 'YZ bay LT rail', '1600543389 / 1600700448',
         'YZ bay LT Rail', 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 29 to 31)', 'Shutdown job'],
        [2, '29.07.2025', 'YZ bay LT rail joint', '1600543390 / 1600700449',
         'YZ bay LT rail joint Thermit welding', 'YZ bay LT rail joints (4 nos.) Thermit welding done. (Column no. 29 to 31)', 'Shutdown job'],
        [3, '30.07.2025', 'YZ bay LT rail', '1600543391 / 1600700540',
         'YZ bay LT Rail', 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 26 to 29)', 'Shutdown job'],
        [4, '30.07.2025', 'YZ bay LT rail joint', '1600543392 / 1600700541',
         'YZ bay LT rail joint Thermit welding', 'YZ bay LT rail joints (4 nos.) Thermit welding done. (Column no. 26 to 29)', 'Shutdown job'],
        [5, '31.07.2025', 'YZ bay LT rail', '1600543394 / 1600700543',
         'YZ bay LT Rail', 'YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 23 to 26)', 'Shutdown job'],
    ]

    for row in sample_rail_rows:
        ws3.append(row)

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 22
    ws3.column_dimensions['D'].width = 26
    ws3.column_dimensions['E'].width = 32
    ws3.column_dimensions['F'].width = 72
    ws3.column_dimensions['G'].width = 18

    ws3.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============== DATA LOADER ==============

def parse_wheel_dataframe(df_raw):
    """
    Normalize a wheel replacement dataframe to internal format:
    Date, Crane, Equipment, Position, Remarks
    """
    df = df_raw.copy()

    if 'S.no.' in df.columns:
        df = df.rename(columns={'S.no.': 'S_No'})

    for col in ['Date', 'Crane', 'Equipment', 'Remarks']:
        if col not in df.columns:
            df[col] = None

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)
    df['Crane'] = df['Crane'].apply(normalize_crane)

    if 'Position' in df.columns:
        df['Position'] = df['Position'].fillna(df.apply(extract_position, axis=1))
    else:
        df['Position'] = df.apply(extract_position, axis=1)

    df['Position'] = df['Position'].apply(
        lambda x: x if x in VALID_POSITIONS else 'Other'
    )

    df['Remarks'] = df['Remarks'].fillna('').astype(str)

    return df

def load_data_from_path(file_path):
    """Load data from a specific Excel file path."""
    if not os.path.exists(file_path):
        return None, None, None

    xls = pd.ExcelFile(file_path)
    sheets = xls.sheet_names

    # Wheel replacement
    if 'LT wheel replacement data' in sheets:
        df_wheels_raw = pd.read_excel(file_path, sheet_name='LT wheel replacement data')
        df_wheels = parse_wheel_dataframe(df_wheels_raw)
    else:
        df_wheels = parse_wheel_dataframe(generate_mock_wheel_data())

    # Hardness data
    if 'Rail Hardness data' in sheets:
        df_hardness_raw = pd.read_excel(file_path, sheet_name='Rail Hardness data', header=None)
        try:
            # Layout: row 0 = empty, row 1 = section labels, row 2 = north HB, row 3 = south HB
            sections = df_hardness_raw.iloc[1, 2:14].astype(str).tolist()
            north = pd.to_numeric(df_hardness_raw.iloc[2, 2:14], errors='coerce').values
            south = pd.to_numeric(df_hardness_raw.iloc[3, 2:14], errors='coerce').values
            df_hardness = pd.DataFrame({'Section': sections, 'North': north, 'South': south})
        except Exception:
            df_hardness = generate_mock_hardness_data()
    else:
        df_hardness = generate_mock_hardness_data()

    # Rail replacement
    if 'Rail Replacement data' in sheets:
        df_rail_raw = pd.read_excel(file_path, sheet_name='Rail Replacement data')
        df_rail = parse_rail_replacement_dataframe(df_rail_raw)
    else:
        df_rail = parse_rail_replacement_dataframe(generate_mock_rail_replacement_data())

    return df_wheels, df_hardness, df_rail

def load_data():
    """Try loading Excel; fall back to mock data if missing or corrupt."""
    if os.path.exists(DATA_PATH):
        try:
            df_wheels, df_hardness, df_rail = load_data_from_path(DATA_PATH)
            return df_wheels, df_hardness, df_rail, True
        except Exception as e:
            print(f"Error loading Excel: {e}. Using mock data.")

    df_wheels = parse_wheel_dataframe(generate_mock_wheel_data())
    return df_wheels, generate_mock_hardness_data(), parse_rail_replacement_dataframe(generate_mock_rail_replacement_data()), False

def reload_data():
    """Reload global data after upload."""
    global df_wheels, df_hardness, df_rail_replacement, using_real_data
    df_wheels, df_hardness, df_rail_replacement, using_real_data = load_data()

# Load data at startup
df_wheels, df_hardness, df_rail_replacement, using_real_data = load_data()

# ============== ROUTES ==============

@app.route('/')
@login_required
def index():
    return render_template('index.html', using_real_data=using_real_data, user=get_current_user())

@app.route('/login', methods=['GET', 'POST'])
def login():
    ensure_db()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('Please enter username and password', 'error')
            return render_template('login.html')

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (username,)
        ).fetchone()
        conn.close()

        if user and bcrypt.check_password_hash(user['password_hash'], password):
            if not user['is_active']:
                flash('Account is disabled', 'error')
                return render_template('login.html')

            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/me')
@login_required
def current_user_api():
    user = get_current_user()
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'role': user['role']
    })

@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    conn = get_db_connection()
    users = conn.execute(
        "SELECT id, username, role, is_active, created_at FROM users ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')

    if not username or not password:
        return jsonify({'success': False, 'error': 'Username and password are required'}), 400

    if role not in ('admin', 'user'):
        return jsonify({'success': False, 'error': 'Role must be admin or user'}), 400

    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            (username, password_hash, role)
        )
        user_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'user_id': user_id}), 201
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'Username already exists'}), 409

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    # Prevent deleting yourself
    if user_id == session['user_id']:
        return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/status')
@login_required
def status():
    user = get_current_user()
    return jsonify({
        'using_real_data': using_real_data,
        'data_path': DATA_PATH,
        'data_path_exists': os.path.exists(DATA_PATH),
        'total_wheel_records': len(df_wheels),
        'hardness_sections': len(df_hardness),
        'user': {
            'id': user['id'],
            'username': user['username'],
            'role': user['role']
        }
    })

@app.route('/api/sample')
@login_required
def download_sample():
    """Download a sample Excel file with data validation."""
    output = generate_sample_excel()
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=SAMPLE_FILENAME
    )

@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    """Handle Excel file upload with in-memory validation and version history."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'errors': ['No file part in the request.'], 'warnings': []}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'errors': ['No file selected.'], 'warnings': []}), 400

    if not allowed_file(file.filename):
        return jsonify({
            'success': False,
            'errors': [f"Invalid file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}."],
            'warnings': []
        }), 400

    user_id = session['user_id']

    try:
        file_bytes = io.BytesIO(file.read())

        try:
            xls = pd.ExcelFile(file_bytes)
            sheets = xls.sheet_names
        except Exception as e:
            return jsonify({
                'success': False,
                'errors': [f"Could not read Excel file: {str(e)}"],
                'warnings': []
            }), 400

        if 'LT wheel replacement data' not in sheets:
            return jsonify({
                'success': False,
                'errors': ["Sheet 'LT wheel replacement data' not found. Please use the sample file format."],
                'warnings': []
            }), 400

        all_errors = []
        all_warnings = []
        all_stats = {}

        df_wheels_raw = pd.read_excel(file_bytes, sheet_name='LT wheel replacement data')
        wheel_validation = validate_wheel_dataframe(df_wheels_raw)
        all_errors.extend(wheel_validation['errors'])
        all_warnings.extend(wheel_validation['warnings'])
        all_stats['wheel_replacement'] = wheel_validation['stats']

        total_wheel_records = wheel_validation['stats'].get('total_rows', 0)

        if 'Rail Hardness data' in sheets:
            try:
                file_bytes.seek(0)
                df_hardness_raw = pd.read_excel(file_bytes, sheet_name='Rail Hardness data', header=None)
                hardness_validation = validate_hardness_dataframe(df_hardness_raw)
                all_errors.extend(hardness_validation['errors'])
                all_warnings.extend(hardness_validation['warnings'])
                all_stats['rail_hardness'] = hardness_validation['stats']
            except Exception as e:
                all_warnings.append(f"Could not validate Rail Hardness data: {str(e)}")

        if 'Rail Replacement data' in sheets:
            try:
                file_bytes.seek(0)
                df_rail_raw = pd.read_excel(file_bytes, sheet_name='Rail Replacement data')
                rail_validation = validate_rail_replacement_dataframe(df_rail_raw)
                all_errors.extend(rail_validation['errors'])
                all_warnings.extend(rail_validation['warnings'])
                all_stats['rail_replacement'] = rail_validation['stats']
            except Exception as e:
                all_warnings.append(f"Could not validate Rail Replacement data: {str(e)}")

        # Only block on truly fatal errors (e.g. completely empty wheel sheet)
        fatal_errors = [e for e in all_errors if 'no data rows' in e.lower()]
        if fatal_errors:
            return jsonify({
                'success': False,
                'errors': fatal_errors,
                'warnings': all_warnings,
                'stats': all_stats
            }), 400

        # Save validated file to disk as active data file
        os.makedirs(DATA_DIR, exist_ok=True)
        file_bytes.seek(0)
        with open(DATA_PATH, 'wb') as f:
            f.write(file_bytes.read())

        # Save as a version in DB
        conn = get_db_connection()
        cursor = conn.cursor()

        # Deactivate previous active version for this user
        cursor.execute(
            "UPDATE excel_versions SET is_active = 0 WHERE user_id = ? AND is_active = 1",
            (user_id,)
        )

        # Calculate version number
        cursor.execute(
            "SELECT COUNT(*) FROM excel_versions WHERE user_id = ?",
            (user_id,)
        )
        version_number = cursor.fetchone()[0] + 1

        original_filename = secure_filename(file.filename)
        stored_filename = f"user_{user_id}_v{version_number}_{original_filename}"

        cursor.execute(
            """INSERT INTO excel_versions
               (user_id, filename, stored_filename, version_number, is_active, total_wheel_records)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (user_id, original_filename, stored_filename, version_number, 1, total_wheel_records)
        )
        version_id = cursor.lastrowid
        conn.commit()
        conn.close()

        # Save actual version file
        file_bytes.seek(0)
        version_path = get_version_storage_path(version_id)
        with open(version_path, 'wb') as f:
            f.write(file_bytes.read())

        # Reload global data
        reload_data()

        return jsonify({
            'success': True,
            'errors': [],
            'warnings': all_warnings,
            'stats': all_stats,
            'using_real_data': using_real_data,
            'total_wheel_records': len(df_wheels),
            'version_id': version_id,
            'version_number': version_number
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'errors': [f"Upload failed: {str(e)}"],
            'warnings': []
        }), 500

@app.route('/api/versions')
@login_required
def list_versions():
    """List Excel versions for the current user (or all if admin)."""
    user = get_current_user()
    conn = get_db_connection()

    if user['role'] == 'admin':
        versions = conn.execute(
            """SELECT v.*, u.username FROM excel_versions v
               JOIN users u ON v.user_id = u.id
               ORDER BY v.created_at DESC"""
        ).fetchall()
    else:
        versions = conn.execute(
            """SELECT v.*, u.username FROM excel_versions v
               JOIN users u ON v.user_id = u.id
               WHERE v.user_id = ?
               ORDER BY v.created_at DESC""",
            (user['id'],)
        ).fetchall()

    conn.close()
    return jsonify([dict(v) for v in versions])

@app.route('/api/versions/<int:version_id>/activate', methods=['POST'])
@login_required
def activate_version(version_id):
    """Activate a previous version and reload data from it."""
    user = get_current_user()
    conn = get_db_connection()

    version = conn.execute(
        "SELECT * FROM excel_versions WHERE id = ?",
        (version_id,)
    ).fetchone()

    if not version:
        conn.close()
        return jsonify({'success': False, 'error': 'Version not found'}), 404

    if user['role'] != 'admin' and version['user_id'] != user['id']:
        conn.close()
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    version_path = get_version_storage_path(version_id)
    if not os.path.exists(version_path):
        conn.close()
        return jsonify({'success': False, 'error': 'Version file not found'}), 404

    # Copy version file to active data path
    shutil.copy2(version_path, DATA_PATH)

    # Update active flags
    conn.execute(
        "UPDATE excel_versions SET is_active = 0 WHERE user_id = ?",
        (version['user_id'],)
    )
    conn.execute(
        "UPDATE excel_versions SET is_active = 1 WHERE id = ?",
        (version_id,)
    )
    conn.commit()
    conn.close()

    # Reload global data
    reload_data()

    return jsonify({
        'success': True,
        'message': f"Version {version['version_number']} activated",
        'version_id': version_id
    })

@app.route('/api/versions/<int:version_id>/download')
@login_required
def download_version(version_id):
    """Download a specific Excel version."""
    user = get_current_user()
    conn = get_db_connection()

    version = conn.execute(
        "SELECT * FROM excel_versions WHERE id = ?",
        (version_id,)
    ).fetchone()
    conn.close()

    if not version:
        return jsonify({'success': False, 'error': 'Version not found'}), 404

    if user['role'] != 'admin' and version['user_id'] != user['id']:
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    version_path = get_version_storage_path(version_id)
    if not os.path.exists(version_path):
        return jsonify({'success': False, 'error': 'Version file not found'}), 404

    return send_file(
        version_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=version['filename']
    )

@app.route('/api/summary')
@login_required
def get_summary():
    df = df_wheels.copy()
    df = df.dropna(subset=['Date'])

    total_replacements = len(df)
    west_count = int(df[df['Crane'].astype(str).str.contains('west', case=False, na=False)].shape[0])
    east_count = int(df[df['Crane'].astype(str).str.contains('East', case=False, na=False)].shape[0])

    df['Month'] = df['Date'].dt.to_period('M')
    monthly = df.groupby('Month').size().reset_index(name='Count')
    monthly['Month_str'] = monthly['Month'].astype(str)
    monthly = monthly.drop(columns=['Month'])

    north_vals = df_hardness['North'].dropna().tolist()
    south_vals = df_hardness['South'].dropna().tolist()
    all_hardness = [float(v) for v in (north_vals + south_vals) if pd.notna(v)]

    avg_north = float(np.mean(north_vals)) if north_vals else 0
    avg_south = float(np.mean(south_vals)) if south_vals else 0
    max_hb = float(max(all_hardness)) if all_hardness else 0
    above_300_north = sum(1 for h in north_vals if h > 300)
    above_300_south = sum(1 for h in south_vals if h > 300)

    df['Year'] = df['Date'].dt.year
    yearly = df.groupby('Year').size().to_dict()

    return jsonify({
        'total_replacements': total_replacements,
        'west_crane_failures': west_count,
        'east_crane_failures': east_count,
        'avg_hardness_north': round(avg_north, 1),
        'avg_hardness_south': round(avg_south, 1),
        'max_hardness': round(max_hb, 1),
        'above_threshold_north': above_300_north,
        'above_threshold_south': above_300_south,
        'threshold': 300,
        'monthly_trend': monthly.to_dict('records'),
        'yearly_trend': {str(k): int(v) for k, v in yearly.items()}
    })

@app.route('/api/hardness-data')
@login_required
def get_hardness_data():
    sections = df_hardness['Section'].astype(str).tolist()
    north = [safe_float(v) for v in df_hardness['North'].values]
    south = [safe_float(v) for v in df_hardness['South'].values]

    return jsonify({
        'columns': sections,
        'north_side': north,
        'south_side': south,
        'threshold': 300,
        'critical_threshold': 400,
        'high_threshold': 350
    })

@app.route('/api/hardness-heatmap')
@login_required
def hardness_heatmap():
    result = []
    for _, row in df_hardness.iterrows():
        section = str(row['Section'])
        for side in ['North', 'South']:
            hb = safe_float(row[side])
            result.append({
                'section': section,
                'side': side,
                'hardness': hb,
                'risk': get_risk_class(hb)
            })
    return jsonify(result)

@app.route('/api/hardness-correlation')
@login_required
def hardness_correlation():
    risk_zones = []
    for _, row in df_hardness.iterrows():
        section = str(row['Section'])
        for side in ['North', 'South']:
            hb = safe_float(row[side])
            if hb is not None:
                risk = get_risk_class(hb)
                action = 'Immediate replacement' if risk == 'Critical' else \
                         'Schedule replacement' if risk == 'High' else \
                         'Monitor monthly' if risk == 'Medium' else 'Normal operation'
                risk_zones.append({
                    'section': f'{side} {section}',
                    'side': side,
                    'hardness': hb,
                    'risk': risk,
                    'recommended_action': action
                })

    all_hbs = [z['hardness'] for z in risk_zones]
    return jsonify({
        'risk_zones': risk_zones,
        'critical_zones': sum(1 for z in risk_zones if z['risk'] == 'Critical'),
        'high_risk_zones': sum(1 for z in risk_zones if z['risk'] == 'High'),
        'medium_risk_zones': sum(1 for z in risk_zones if z['risk'] == 'Medium'),
        'normal_zones': sum(1 for z in risk_zones if z['risk'] == 'Normal'),
        'max_hardness': round(max(all_hbs), 1) if all_hbs else 0,
        'avg_hardness': round(float(np.mean(all_hbs)), 1) if all_hbs else 0,
        'recommendation': 'Execute complete rail replacement program across critical zones'
    })

@app.route('/api/wheel-failure-analysis')
@login_required
def wheel_failure_analysis():
    df = df_wheels.copy()
    df = df.dropna(subset=['Date'])

    equipment_counts = df['Equipment'].value_counts().head(10).to_dict()

    if 'Position' in df.columns:
        positions = df['Position'].fillna('Other').tolist()
    else:
        positions = []
        for equip in df['Equipment'].astype(str):
            pos = 'Other'
            for p in ['SW', 'NE', 'NW', 'SE']:
                if p in equip:
                    pos = p
                    break
            positions.append(pos)

    position_counts = pd.Series(positions).value_counts().to_dict()

    job_counts = df['Remarks'].value_counts().head(8).to_dict()

    df['Month'] = df['Date'].dt.to_period('M')
    monthly = df.groupby('Month').size().to_dict()

    return jsonify({
        'equipment_failures': {str(k): int(v) for k, v in equipment_counts.items()},
        'position_failures': {str(k): int(v) for k, v in position_counts.items()},
        'job_types': {str(k): int(v) for k, v in job_counts.items()},
        'total_by_month': {str(k): int(v) for k, v in monthly.items()}
    })

@app.route('/api/failure-distribution')
@login_required
def failure_distribution():
    df = df_wheels.copy()
    df = df.dropna(subset=['Date'])

    crane_counts = df['Crane'].value_counts().to_dict()

    df['Quarter'] = df['Date'].dt.to_period('Q')
    quarterly = df.groupby('Quarter').size().to_dict()

    df['DayOfWeek'] = df['Date'].dt.day_name()
    dow = df['DayOfWeek'].value_counts().to_dict()

    severity = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
    for remark in df['Remarks'].astype(str):
        r = remark.lower()
        if any(x in r for x in ['crack', 'hot axle', 'bearing failure', 'breakdown']):
            severity['Critical'] += 1
        elif any(x in r for x in ['damaged', 'worn out', 'broken']):
            severity['High'] += 1
        elif any(x in r for x in ['wear', 'limit', 'reduced']):
            severity['Medium'] += 1
        else:
            severity['Low'] += 1

    return jsonify({
        'by_crane': {str(k): int(v) for k, v in crane_counts.items()},
        'by_quarter': {str(k): int(v) for k, v in quarterly.items()},
        'by_day_of_week': dow,
        'by_severity': severity
    })

@app.route('/api/rail-replacement')
@login_required
def rail_replacement():
    df = df_rail_replacement.copy()
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    records = df.fillna('').to_dict('records')
    for r in records:
        for k in r:
            if isinstance(r[k], pd.Timestamp):
                r[k] = r[k].strftime('%Y-%m-%d')
    return jsonify({
        'records': records,
        'total_pieces': int(df['Qty_Pieces'].sum()) if 'Qty_Pieces' in df.columns else 0,
        'total_events': len(df)
    })

@app.route('/api/predict/<int:months_ahead>')
@login_required
def predict_failures(months_ahead):
    df = df_wheels.copy()
    df = df.dropna(subset=['Date'])

    if len(df) < 2:
        return jsonify({'error': 'Insufficient data'})

    df['Days'] = (df['Date'] - df['Date'].min()).dt.days
    time_series = df.groupby('Days').size().cumsum().reset_index()
    time_series.columns = ['Days', 'Cumulative_Failures']

    X = time_series['Days'].values.reshape(-1, 1)
    y = time_series['Cumulative_Failures'].values

    lin_model = LinearRegression()
    lin_model.fit(X, y)

    poly = PolynomialFeatures(degree=2)
    X_poly = poly.fit_transform(X)
    poly_model = LinearRegression()
    poly_model.fit(X_poly, y)

    last_day = int(time_series['Days'].max())
    future_days = [last_day + (30 * i) for i in range(1, months_ahead + 1)]

    future_linear = lin_model.predict(np.array(future_days).reshape(-1, 1))
    future_poly = poly_model.predict(poly.transform(np.array(future_days).reshape(-1, 1)))

    all_hb = []
    for col in ['North', 'South']:
        all_hb.extend([float(v) for v in df_hardness[col].values if pd.notna(v)])
    avg_hb = float(np.mean(all_hb)) if all_hb else 300
    hardness_risk = max(0, (avg_hb - 300) / 100)

    adjusted_linear = future_linear * (1 + hardness_risk)
    adjusted_poly = np.maximum.accumulate(future_poly * (1 + hardness_risk))

    current_failures = int(y[-1])
    pred_3mo = int(adjusted_poly[2] - current_failures) if months_ahead >= 3 else 0
    pred_6mo = int(adjusted_poly[5] - current_failures) if months_ahead >= 6 else 0
    monthly_rate = (adjusted_poly[-1] - current_failures) / months_ahead

    if avg_hb > 400:
        recommendation = 'URGENT: Critical hardness levels - immediate rail replacement required'
    elif avg_hb > 350:
        recommendation = 'HIGH PRIORITY: Schedule comprehensive rail replacement program'
    elif avg_hb > 300:
        recommendation = 'MODERATE: Monitor hardness monthly and plan targeted replacement'
    else:
        recommendation = 'NORMAL: Continue standard inspection intervals'

    return jsonify({
        'current_failures': current_failures,
        'predicted_failures_next_3mo': pred_3mo,
        'predicted_failures_next_6mo': pred_6mo,
        'avg_hardness': round(avg_hb, 1),
        'hardness_risk_factor': round(hardness_risk, 2),
        'monthly_failure_rate': round(monthly_rate, 1),
        'recommendation': recommendation,
        'months_ahead': months_ahead,
        'future_predictions': [
            {
                'month_index': i + 1,
                'linear': round(float(adjusted_linear[i] - current_failures), 1),
                'polynomial': round(float(adjusted_poly[i] - current_failures), 1)
            }
            for i in range(months_ahead)
        ]
    })

@app.route('/api/scatter-hardness-failures')
@login_required
def scatter_hardness_failures():
    """Simulated correlation: hardness vs failure count by section"""
    df = df_wheels.copy()
    points = []
    for _, row in df_hardness.iterrows():
        section = str(row['Section'])
        north_hb = safe_float(row['North']) or 300
        south_hb = safe_float(row['South']) or 300
        avg_hb = (north_hb + south_hb) / 2

        base_failures = max(0, (avg_hb - 280) / 10 + np.random.normal(0, 1))
        points.append({
            'section': section,
            'avg_hardness': round(avg_hb, 1),
            'estimated_failures': round(base_failures, 1),
            'north': north_hb,
            'south': south_hb
        })
    return jsonify(points)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV', 'production') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
