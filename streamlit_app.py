import streamlit as st
import sqlite3
import bcrypt
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import os
import io
import re
import shutil
from datetime import datetime, timedelta
import warnings

warnings.filterwarnings("ignore")

# ─────────────────────── CONFIGURATION ───────────────────────

DATA_DIR = "data"
DATA_PATH = os.path.join(DATA_DIR, "LT Wheel replacement data.xlsx")
SAMPLE_FILENAME = "LT Wheel replacement data.xlsx"
DB_PATH = os.path.join(DATA_DIR, "crane_analytics.db")
VALID_POSITIONS = {"SW", "SE", "NE", "NW"}
REQUIRED_WHEEL_COLUMNS = {"Date", "Crane", "Equipment", "Remarks"}

st.set_page_config(
    page_title="YZ Bay | Crane LT Wheel Analytics",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────── DATABASE ───────────────────────

def get_db_connection():
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS excel_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            version_number INTEGER NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            total_wheel_records INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
        )
    """)
    cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
    if cursor.fetchone()[0] == 0:
        admin_hash = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode("utf-8")
        cursor.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            ("admin", admin_hash, "admin"),
        )
    conn.commit()
    conn.close()


# ─────────────────────── AUTH HELPERS ───────────────────────

def check_password(password: str, hash_str: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode("utf-8"), hash_str.encode("utf-8"))
    except Exception:
        return False


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


# ─────────────────────── HELPERS ───────────────────────

def safe_float(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def get_risk_class(hb):
    if hb is None:
        return "Unknown"
    if hb > 400:
        return "Critical"
    if hb > 350:
        return "High"
    if hb > 300:
        return "Medium"
    return "Normal"


def normalize_crane(crane):
    if pd.isna(crane):
        return "UNKNOWN"
    c = str(crane).lower()
    if "west" in c:
        return "LT WEST"
    if "east" in c:
        return "LT EAST"
    return str(crane).strip()


def extract_position(row):
    for col in ["Equipment", "Job Description"]:
        if col in row and pd.notna(row[col]):
            text = str(row[col]).upper()
            for pos in VALID_POSITIONS:
                if re.search(r"\b" + pos + r"\b", text):
                    return pos
    return "Other"


def get_version_storage_path(version_id):
    return os.path.join(DATA_DIR, f"version_{version_id}.xlsx")


# ─────────────────────── MOCK DATA ───────────────────────

def generate_mock_wheel_data():
    np.random.seed(42)
    cranes = ["YZ bay west crane", "YZ bay East crane"]
    positions = ["SW", "NE", "NW", "SE"]
    wheel_types = ["drive wheel", "idle wheel"]
    remarks_list = [
        "Shutdown job",
        "Breakdown job",
        "Planned maintenance",
        "Inspection replacement",
    ]
    data = []
    start_date = datetime(2024, 1, 1)
    end_date = datetime(2025, 9, 30)
    for i in range(87):
        crane = np.random.choice(cranes, p=[0.58, 0.42])
        pos = np.random.choice(positions)
        wheel_type = np.random.choice(wheel_types)
        equipment = f"LT {pos} {wheel_type}"
        days_offset = int(np.random.beta(2, 1.2) * (end_date - start_date).days)
        date = start_date + timedelta(days=days_offset)
        remark = np.random.choice(remarks_list)
        job_desc = (
            f"{crane} LT {pos} {wheel_type} replacement done. "
            f"(Wheel collar {'reduced' if remark == 'Shutdown job' else 'broken'})"
        )
        data.append(
            {
                "S.no.": i + 1,
                "Date": date,
                "Crane": crane,
                "Equipment": equipment,
                "Job Description": job_desc,
                "Remarks": remark,
            }
        )
    df = pd.DataFrame(data).sort_values("Date").reset_index(drop=True)
    df["S.no."] = range(1, len(df) + 1)
    return df


def generate_mock_hardness_data():
    np.random.seed(7)
    sections = [
        "21-22", "22-23", "23-24", "24-25", "25-26", "26-27",
        "27-28", "28-29", "29-30", "30-31", "31-32", "End portion 32-33",
    ]
    north = [270, 270, 330, 466, 374, 412, 420, 380, 380, 360, 320, 270]
    north = [h + np.random.randint(-8, 9) for h in north]
    south = [270, 270, 330, 395, 390, 370, 380, 400, 355, 380, 360, 270]
    south = [h + np.random.randint(-7, 8) for h in south]
    return pd.DataFrame({"Section": sections, "North": north, "South": south})


def generate_mock_rail_replacement_data():
    data = [
        {
            "S.no.": 1,
            "Date": "29.07.2025",
            "Crane": "YZ bay LT rail",
            "Equipment": "YZ bay LT Rail",
            "Job Description": "YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 29 to 31)",
            "Remarks": "Shutdown job",
        },
        {
            "S.no.": 2,
            "Date": "29.07.2025",
            "Crane": "YZ bay LT rail joint",
            "Equipment": "YZ bay LT rail joint Thermit welding",
            "Job Description": "YZ bay LT rail joints (4 nos.) Thermit welding done. (Column no. 29 to 31)",
            "Remarks": "Shutdown job",
        },
        {
            "S.no.": 3,
            "Date": "30.07.2025",
            "Crane": "YZ bay LT rail",
            "Equipment": "YZ bay LT Rail",
            "Job Description": "YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 26 to 29)",
            "Remarks": "Shutdown job",
        },
    ]
    return pd.DataFrame(data)


# ─────────────────────── PARSERS ───────────────────────

def extract_rail_qty(job_description):
    if pd.isna(job_description):
        return 0
    match = re.search(r"(\d+)\s*no\.?", str(job_description), re.IGNORECASE)
    return int(match.group(1)) if match else 0


def extract_rail_section(job_description):
    if pd.isna(job_description):
        return "Unknown"
    text = str(job_description)
    match = re.search(r"Column no\.\s*(\d+)\s*to\s*(\d+)", text, re.IGNORECASE)
    if match:
        return f"Column {match.group(1)} to {match.group(2)}"
    match = re.search(r"Column no\.\s*(\d+)", text, re.IGNORECASE)
    if match:
        return f"Column {match.group(1)}"
    return "Unknown"


def extract_rail_reason(equipment, job_description):
    text = " ".join([str(equipment), str(job_description)]).lower()
    if "thermit" in text or "welding" in text:
        return "Thermit welding"
    if "replacement" in text:
        return "Rail pieces replacement"
    return "Maintenance"


def parse_rail_replacement_dataframe(df_raw):
    df = df_raw.copy()
    df.columns = [str(c).strip() for c in df.columns]
    old_cols = {"Date", "Section", "Side", "Qty_Pieces", "Reason"}
    if old_cols.issubset(set(df.columns)):
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True)
        df["Qty_Pieces"] = pd.to_numeric(df["Qty_Pieces"], errors="coerce").fillna(0).astype(int)
        return df[["Date", "Section", "Side", "Qty_Pieces", "Reason"]]
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True)
    df["Qty_Pieces"] = df.apply(lambda r: extract_rail_qty(r.get("Job Description")), axis=1)
    df["Section"] = df.apply(lambda r: extract_rail_section(r.get("Job Description")), axis=1)
    df["Reason"] = df.apply(
        lambda r: extract_rail_reason(r.get("Equipment", ""), r.get("Job Description", "")),
        axis=1,
    )
    df["Side"] = "Both"
    return df[["Date", "Section", "Side", "Qty_Pieces", "Reason"]]


def parse_wheel_dataframe(df_raw):
    df = df_raw.copy()
    if "S.no." in df.columns:
        df = df.rename(columns={"S.no.": "S_No"})
    for col in ["Date", "Crane", "Equipment", "Remarks"]:
        if col not in df.columns:
            df[col] = None
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True)
    df["Crane"] = df["Crane"].apply(normalize_crane)
    if "Position" in df.columns:
        df["Position"] = df["Position"].fillna(df.apply(extract_position, axis=1))
    else:
        df["Position"] = df.apply(extract_position, axis=1)
    df["Position"] = df["Position"].apply(lambda x: x if x in VALID_POSITIONS else "Other")
    df["Remarks"] = df["Remarks"].fillna("").astype(str)
    return df


# ─────────────────────── DATA LOADER ───────────────────────

def load_data_from_path(file_path):
    xls = pd.ExcelFile(file_path)
    sheets = xls.sheet_names

    if "LT wheel replacement data" in sheets:
        df_wheels = parse_wheel_dataframe(
            pd.read_excel(file_path, sheet_name="LT wheel replacement data")
        )
    else:
        df_wheels = parse_wheel_dataframe(generate_mock_wheel_data())

    if "Rail Hardness data" in sheets:
        df_h_raw = pd.read_excel(file_path, sheet_name="Rail Hardness data", header=None)
        try:
            sections = df_h_raw.iloc[1, 2:14].astype(str).tolist()
            north = pd.to_numeric(df_h_raw.iloc[2, 2:14], errors="coerce").values
            south = pd.to_numeric(df_h_raw.iloc[3, 2:14], errors="coerce").values
            df_hardness = pd.DataFrame({"Section": sections, "North": north, "South": south})
        except Exception:
            df_hardness = generate_mock_hardness_data()
    else:
        df_hardness = generate_mock_hardness_data()

    if "Rail Replacement data" in sheets:
        df_rail = parse_rail_replacement_dataframe(
            pd.read_excel(file_path, sheet_name="Rail Replacement data")
        )
    else:
        df_rail = parse_rail_replacement_dataframe(generate_mock_rail_replacement_data())

    return df_wheels, df_hardness, df_rail


@st.cache_data
def load_data():
    if os.path.exists(DATA_PATH):
        try:
            df_wheels, df_hardness, df_rail = load_data_from_path(DATA_PATH)
            return df_wheels, df_hardness, df_rail, True
        except Exception as e:
            st.warning(f"Error loading Excel: {e}. Using demo data.")
    df_wheels = parse_wheel_dataframe(generate_mock_wheel_data())
    return (
        df_wheels,
        generate_mock_hardness_data(),
        parse_rail_replacement_dataframe(generate_mock_rail_replacement_data()),
        False,
    )


# ─────────────────────── SAMPLE EXCEL ───────────────────────

def generate_sample_excel():
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "LT wheel replacement data"
    headers1 = ["S.no.", "Date", "Crane", "Equipment", "Job Description", "Remarks"]
    ws1.append(headers1)
    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True)
    for col_num, _ in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws1.append([
        1, "02.04.2024", "YZ bay west crane", "LT SW drive wheel",
        "YZ bay west crane LT SW drive wheel replacement done. (Wheel collar reduced)", "Shutdown job",
    ])
    ws1.append([
        2, "06.04.2024", "YZ bay East crane", "LT NE idle wheel",
        "YZ bay east crane LT NE idle wheel replacement done. (Wheel collar broken)", "Breakdown job",
    ])

    crane_validation = DataValidation(
        type="list",
        formula1='"YZ bay west crane,YZ bay East crane"',
        allow_blank=False,
    )
    ws1.add_data_validation(crane_validation)
    crane_validation.add("C2:C1000")

    ws2 = wb.create_sheet("Rail Hardness data")
    ws2.append([
        "", "Column no.", "21-22", "22-23", "23-24", "24-25", "25-26",
        "26-27", "27-28", "28-29", "29-30", "30-31", "31-32", "End portion 32-33",
    ])
    ws2.append(["", "North Side Z bay Rail hardness (HB)", 270, 270, 330, 466, 374, 412, 420, 380, 380, 360, 320, 270])
    ws2.append(["", "South Side Y bay Rail hardness (HB)", 270, 270, 330, 395, 390, 370, 380, 400, 355, 380, 360, 270])

    ws3 = wb.create_sheet("Rail Replacement data")
    ws3.append(["S.no.", "Date", "Crane", "Notification no.", "Equipment", "Job Description", "Remarks"])
    ws3.append([
        1, "29.07.2025", "YZ bay LT rail", "1600543389 / 1600700448",
        "YZ bay LT Rail",
        "YZ bay LT 6 nos. of Rail pieces replacement done. (12 m each) (Column no. 29 to 31)",
        "Shutdown job",
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────── LOGIN PAGE ───────────────────────

def show_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("## 🏗️ YZ Bay Crane Analytics")
        st.markdown("*Predictive maintenance dashboard for crane LT wheel failure reduction*")
        st.divider()
        with st.form("login_form"):
            st.subheader("Sign In")
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            submitted = st.form_submit_button("Sign In", use_container_width=True, type="primary")
            if submitted:
                if not username or not password:
                    st.error("Please enter username and password")
                else:
                    conn = get_db_connection()
                    user = conn.execute(
                        "SELECT * FROM users WHERE username = ?", (username,)
                    ).fetchone()
                    conn.close()
                    if user and check_password(password, user["password_hash"]):
                        if not user["is_active"]:
                            st.error("Account is disabled")
                        else:
                            st.session_state.logged_in = True
                            st.session_state.user_id = user["id"]
                            st.session_state.username = user["username"]
                            st.session_state.role = user["role"]
                            st.rerun()
                    else:
                        st.error("Invalid username or password")


# ─────────────────────── DASHBOARD ───────────────────────

def show_dashboard(df_wheels, df_hardness, df_rail, using_real_data):
    st.title("🏗️ YZ Bay — Crane LT Wheel Failure Dashboard")
    if not using_real_data:
        st.info("ℹ️ Displaying demo data. Upload your Excel file in **Upload Data** to see real analytics.")

    df = df_wheels.dropna(subset=["Date"]).copy()

    # KPIs
    total = len(df)
    west = df[df["Crane"].str.contains("WEST", case=False, na=False)].shape[0]
    east = df[df["Crane"].str.contains("EAST", case=False, na=False)].shape[0]
    all_hb = [safe_float(v) for v in list(df_hardness["North"]) + list(df_hardness["South"]) if safe_float(v)]
    avg_hb = round(np.mean(all_hb), 1) if all_hb else 0
    max_hb = round(max(all_hb), 1) if all_hb else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Total Replacements", total)
    k2.metric("LT West Crane", west)
    k3.metric("LT East Crane", east)
    k4.metric("Avg Rail Hardness (HB)", avg_hb)
    k5.metric(
        "Max Rail Hardness (HB)",
        max_hb,
        delta="⚠️ Critical" if max_hb > 400 else None,
        delta_color="inverse",
    )

    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Monthly Wheel Replacements")
        df["Month"] = df["Date"].dt.to_period("M").astype(str)
        monthly = df.groupby("Month").size().reset_index(name="Count")
        fig = px.bar(monthly, x="Month", y="Count", color="Count",
                     color_continuous_scale="blues", labels={"Month": "", "Count": "Replacements"})
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Replacements by Crane")
        crane_counts = df["Crane"].value_counts().reset_index()
        crane_counts.columns = ["Crane", "Count"]
        fig = px.pie(crane_counts, names="Crane", values="Count",
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        st.subheader("Rail Hardness by Section")
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df_hardness["Section"], y=df_hardness["North"],
                                 mode="lines+markers", name="North", line=dict(color="#2563eb")))
        fig.add_trace(go.Scatter(x=df_hardness["Section"], y=df_hardness["South"],
                                 mode="lines+markers", name="South", line=dict(color="#ec4899")))
        fig.add_hline(y=300, line_dash="dash", line_color="orange", annotation_text="Threshold 300 HB")
        fig.add_hline(y=400, line_dash="dash", line_color="red", annotation_text="Critical 400 HB")
        fig.update_layout(xaxis_tickangle=-45, margin=dict(t=10, b=10), legend=dict(x=0, y=1))
        st.plotly_chart(fig, use_container_width=True)

    with col4:
        st.subheader("Failures by Position")
        pos_counts = df["Position"].value_counts().reset_index()
        pos_counts.columns = ["Position", "Count"]
        fig = px.bar(pos_counts, x="Position", y="Count",
                     color="Position", color_discrete_sequence=px.colors.qualitative.Pastel)
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)


# ─────────────────────── HARDNESS ANALYSIS ───────────────────────

def show_hardness(df_hardness):
    st.title("🔬 Rail Hardness Analysis")

    risk_data = []
    for _, row in df_hardness.iterrows():
        for side in ["North", "South"]:
            hb = safe_float(row[side])
            risk = get_risk_class(hb)
            action = (
                "Immediate replacement" if risk == "Critical"
                else "Schedule replacement" if risk == "High"
                else "Monitor monthly" if risk == "Medium"
                else "Normal operation"
            )
            risk_data.append({"Section": row["Section"], "Side": side,
                               "Hardness (HB)": hb, "Risk": risk, "Action": action})
    risk_df = pd.DataFrame(risk_data)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Critical Zones", risk_df[risk_df["Risk"] == "Critical"].shape[0])
    c2.metric("High Risk Zones", risk_df[risk_df["Risk"] == "High"].shape[0])
    c3.metric("Medium Zones", risk_df[risk_df["Risk"] == "Medium"].shape[0])
    c4.metric("Normal Zones", risk_df[risk_df["Risk"] == "Normal"].shape[0])

    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("North vs South Hardness per Section")
        fig = go.Figure()
        fig.add_trace(go.Bar(name="North", x=df_hardness["Section"], y=df_hardness["North"],
                             marker_color="#2563eb"))
        fig.add_trace(go.Bar(name="South", x=df_hardness["Section"], y=df_hardness["South"],
                             marker_color="#ec4899"))
        fig.add_hline(y=300, line_dash="dash", line_color="orange", annotation_text="Normal (300)")
        fig.add_hline(y=350, line_dash="dash", line_color="red", annotation_text="High (350)")
        fig.add_hline(y=400, line_dash="dash", line_color="darkred", annotation_text="Critical (400)")
        fig.update_layout(barmode="group", xaxis_tickangle=-45, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Risk Distribution")
        risk_counts = risk_df["Risk"].value_counts().reset_index()
        risk_counts.columns = ["Risk", "Count"]
        color_map = {
            "Normal": "#10b981", "Medium": "#f59e0b",
            "High": "#ef4444", "Critical": "#7c3aed", "Unknown": "#64748b",
        }
        fig = px.pie(risk_counts, names="Risk", values="Count",
                     color="Risk", color_discrete_map=color_map)
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    st.subheader("Hardness Heatmap")
    pivot = risk_df.pivot(index="Side", columns="Section", values="Hardness (HB)")
    fig = px.imshow(pivot, color_continuous_scale="RdYlGn_r", zmin=250, zmax=480,
                    aspect="auto", labels=dict(color="Hardness (HB)"))
    fig.update_layout(margin=dict(t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Risk Zone Details")
    color_map_style = {
        "Critical": "background-color: #fce4ec",
        "High": "background-color: #fff3e0",
        "Medium": "background-color: #fff8e1",
        "Normal": "background-color: #e8f5e9",
    }
    styled = risk_df.style.map(lambda v: color_map_style.get(v, ""), subset=["Risk"])
    st.dataframe(styled, use_container_width=True)


# ─────────────────────── WHEEL FAILURE ANALYSIS ───────────────────────

def show_wheel_failure(df_wheels):
    st.title("⚙️ Wheel Failure Analysis")
    df = df_wheels.dropna(subset=["Date"]).copy()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Failures by Equipment Type")
        equip = df["Equipment"].value_counts().head(10).reset_index()
        equip.columns = ["Equipment", "Count"]
        fig = px.bar(equip, x="Count", y="Equipment", orientation="h",
                     color="Count", color_continuous_scale="blues")
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10), yaxis=dict(autorange="reversed"))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Failures by Position (SW / SE / NE / NW)")
        pos = df["Position"].value_counts().reset_index()
        pos.columns = ["Position", "Count"]
        fig = px.pie(pos, names="Position", values="Count",
                     color_discrete_sequence=px.colors.qualitative.Set3)
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        st.subheader("Failures by Remark Type")
        remarks = df["Remarks"].value_counts().reset_index()
        remarks.columns = ["Remark", "Count"]
        fig = px.bar(remarks, x="Remark", y="Count", color="Remark",
                     color_discrete_sequence=px.colors.qualitative.Pastel)
        fig.update_layout(showlegend=False, xaxis_tickangle=-30, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col4:
        st.subheader("Monthly Trend")
        df["Month"] = df["Date"].dt.to_period("M").astype(str)
        monthly = df.groupby("Month").size().reset_index(name="Count")
        fig = px.line(monthly, x="Month", y="Count", markers=True, line_shape="spline")
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    st.subheader("Raw Records")
    display_cols = [c for c in ["Date", "Crane", "Equipment", "Position", "Remarks"] if c in df.columns]
    st.dataframe(df[display_cols].sort_values("Date", ascending=False), use_container_width=True)


# ─────────────────────── FAILURE DISTRIBUTION ───────────────────────

def show_failure_distribution(df_wheels):
    st.title("📊 Failure Distribution")
    df = df_wheels.dropna(subset=["Date"]).copy()

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("By Crane")
        crane_counts = df["Crane"].value_counts().reset_index()
        crane_counts.columns = ["Crane", "Count"]
        fig = px.pie(crane_counts, names="Crane", values="Count",
                     color_discrete_sequence=px.colors.qualitative.Set1)
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Quarterly Trend")
        df["Quarter"] = df["Date"].dt.to_period("Q").astype(str)
        quarterly = df.groupby("Quarter").size().reset_index(name="Count")
        fig = px.bar(quarterly, x="Quarter", y="Count", color="Count",
                     color_continuous_scale="viridis")
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        st.subheader("By Day of Week")
        df["DayOfWeek"] = df["Date"].dt.day_name()
        day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        dow = df["DayOfWeek"].value_counts().reindex(day_order, fill_value=0).reset_index()
        dow.columns = ["Day", "Count"]
        fig = px.bar(dow, x="Day", y="Count", color="Day",
                     color_discrete_sequence=px.colors.qualitative.Pastel)
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

    with col4:
        st.subheader("By Severity")
        severity = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        for remark in df["Remarks"].astype(str):
            r = remark.lower()
            if any(x in r for x in ["crack", "hot axle", "bearing failure", "breakdown"]):
                severity["Critical"] += 1
            elif any(x in r for x in ["damaged", "worn out", "broken"]):
                severity["High"] += 1
            elif any(x in r for x in ["wear", "limit", "reduced"]):
                severity["Medium"] += 1
            else:
                severity["Low"] += 1
        sev_df = pd.DataFrame(list(severity.items()), columns=["Severity", "Count"])
        color_map = {"Critical": "#ef4444", "High": "#f97316", "Medium": "#f59e0b", "Low": "#10b981"}
        fig = px.bar(sev_df, x="Severity", y="Count", color="Severity",
                     color_discrete_map=color_map)
        fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)


# ─────────────────────── RAIL REPLACEMENT ───────────────────────

def show_rail_replacement(df_rail):
    st.title("🛤️ Rail Replacement Log")
    df = df_rail.copy()

    total_pieces = int(df["Qty_Pieces"].sum()) if "Qty_Pieces" in df.columns else 0
    total_events = len(df)

    c1, c2 = st.columns(2)
    c1.metric("Total Replacement Events", total_events)
    c2.metric("Total Rail Pieces Replaced", total_pieces)

    st.divider()

    if "Reason" in df.columns and "Section" in df.columns:
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("By Reason")
            reason_counts = df["Reason"].value_counts().reset_index()
            reason_counts.columns = ["Reason", "Count"]
            fig = px.pie(reason_counts, names="Reason", values="Count",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_layout(margin=dict(t=10, b=10))
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.subheader("Rail Pieces by Section")
            section_qty = df.groupby("Section")["Qty_Pieces"].sum().reset_index()
            section_qty.columns = ["Section", "Pieces"]
            fig = px.bar(section_qty, x="Section", y="Pieces",
                         color="Pieces", color_continuous_scale="reds")
            fig.update_layout(showlegend=False, margin=dict(t=10, b=10))
            st.plotly_chart(fig, use_container_width=True)

    st.subheader("Replacement Records")
    st.dataframe(df, use_container_width=True)


# ─────────────────────── PREDICTIONS ───────────────────────

def show_predictions(df_wheels, df_hardness):
    st.title("🔮 Failure Predictions")
    df = df_wheels.dropna(subset=["Date"]).copy()

    months_ahead = st.slider("Months to predict ahead", min_value=3, max_value=24, value=12, step=1)

    if len(df) < 2:
        st.error("Insufficient data for prediction.")
        return

    df["Days"] = (df["Date"] - df["Date"].min()).dt.days
    time_series = df.groupby("Days").size().cumsum().reset_index()
    time_series.columns = ["Days", "Cumulative_Failures"]

    X = time_series["Days"].values.reshape(-1, 1)
    y_vals = time_series["Cumulative_Failures"].values

    lin_model = LinearRegression()
    lin_model.fit(X, y_vals)

    poly = PolynomialFeatures(degree=2)
    X_poly = poly.fit_transform(X)
    poly_model = LinearRegression()
    poly_model.fit(X_poly, y_vals)

    last_day = int(time_series["Days"].max())
    future_days = [last_day + (30 * i) for i in range(1, months_ahead + 1)]

    future_linear = lin_model.predict(np.array(future_days).reshape(-1, 1))
    future_poly = poly_model.predict(poly.transform(np.array(future_days).reshape(-1, 1)))

    all_hb = [safe_float(v) for v in list(df_hardness["North"]) + list(df_hardness["South"]) if safe_float(v)]
    avg_hb = float(np.mean(all_hb)) if all_hb else 300
    hardness_risk = max(0, (avg_hb - 300) / 100)

    adjusted_linear = future_linear * (1 + hardness_risk)
    adjusted_poly = np.maximum.accumulate(future_poly * (1 + hardness_risk))

    current_failures = int(y_vals[-1])
    pred_3mo = int(adjusted_poly[2] - current_failures) if months_ahead >= 3 else 0
    pred_6mo = int(adjusted_poly[5] - current_failures) if months_ahead >= 6 else 0
    monthly_rate = round((adjusted_poly[-1] - current_failures) / months_ahead, 1)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Current Total Failures", current_failures)
    c2.metric("Next 3 Months (predicted)", pred_3mo)
    c3.metric("Next 6 Months (predicted)", pred_6mo)
    c4.metric("Avg Monthly Rate", monthly_rate)

    if avg_hb > 400:
        st.error("🚨 URGENT: Critical hardness levels — immediate rail replacement required")
    elif avg_hb > 350:
        st.warning("⚠️ HIGH PRIORITY: Schedule comprehensive rail replacement program")
    elif avg_hb > 300:
        st.warning("📋 MODERATE: Monitor hardness monthly and plan targeted replacement")
    else:
        st.success("✅ NORMAL: Continue standard inspection intervals")

    st.divider()
    st.subheader("Cumulative Failure Forecast")

    future_months = [f"Month +{i + 1}" for i in range(months_ahead)]
    pred_df = pd.DataFrame({
        "Month": future_months,
        "Linear Forecast": [round(float(adjusted_linear[i] - current_failures), 1) for i in range(months_ahead)],
        "Polynomial Forecast": [round(float(adjusted_poly[i] - current_failures), 1) for i in range(months_ahead)],
    })

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=pred_df["Month"], y=pred_df["Linear Forecast"],
                             name="Linear", mode="lines+markers",
                             line=dict(color="#2563eb", dash="dot")))
    fig.add_trace(go.Scatter(x=pred_df["Month"], y=pred_df["Polynomial Forecast"],
                             name="Polynomial", mode="lines+markers",
                             line=dict(color="#ec4899")))
    fig.update_layout(xaxis_tickangle=-45, margin=dict(t=10, b=10),
                      yaxis_title="Additional Failures", legend=dict(x=0, y=1))
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Prediction Table")
    st.dataframe(pred_df, use_container_width=True)


# ─────────────────────── UPLOAD DATA ───────────────────────

def show_upload():
    st.title("📤 Upload Maintenance Data")
    st.info(
        "Upload an Excel file with 3 sheets: "
        "**LT wheel replacement data**, **Rail Hardness data**, **Rail Replacement data**."
    )

    col1, col2 = st.columns([2, 1])

    with col1:
        uploaded = st.file_uploader("Choose an Excel file (.xlsx / .xls)", type=["xlsx", "xls"])
        if uploaded:
            try:
                file_bytes = io.BytesIO(uploaded.read())
                xls = pd.ExcelFile(file_bytes)
                sheets = xls.sheet_names

                if "LT wheel replacement data" not in sheets:
                    st.error("Sheet 'LT wheel replacement data' not found. Please use the sample file format.")
                    return

                file_bytes.seek(0)
                df_wheels_raw = pd.read_excel(file_bytes, sheet_name="LT wheel replacement data")
                if df_wheels_raw.dropna(how="all").empty:
                    st.error("Wheel replacement sheet has no data rows.")
                    return

                os.makedirs(DATA_DIR, exist_ok=True)
                file_bytes.seek(0)
                with open(DATA_PATH, "wb") as f:
                    f.write(file_bytes.read())

                user_id = st.session_state.user_id
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE excel_versions SET is_active = 0 WHERE user_id = ? AND is_active = 1",
                    (user_id,),
                )
                cursor.execute("SELECT COUNT(*) FROM excel_versions WHERE user_id = ?", (user_id,))
                version_number = cursor.fetchone()[0] + 1
                stored_filename = f"user_{user_id}_v{version_number}_{uploaded.name}"
                cursor.execute(
                    "INSERT INTO excel_versions (user_id, filename, stored_filename, version_number, is_active, total_wheel_records) VALUES (?, ?, ?, ?, ?, ?)",
                    (user_id, uploaded.name, stored_filename, version_number, 1,
                     len(df_wheels_raw.dropna(how="all"))),
                )
                version_id = cursor.lastrowid
                conn.commit()
                conn.close()

                file_bytes.seek(0)
                version_path = get_version_storage_path(version_id)
                with open(version_path, "wb") as f:
                    f.write(file_bytes.read())

                st.cache_data.clear()
                st.success(f"✅ File uploaded as version {version_number}! Navigate to Dashboard to see updated data.")
                st.balloons()
            except Exception as e:
                st.error(f"Upload failed: {e}")

    with col2:
        st.subheader("📥 Sample Template")
        st.write("Download a ready-to-fill Excel template.")
        sample = generate_sample_excel()
        st.download_button(
            label="⬇️ Download Sample File",
            data=sample,
            file_name=SAMPLE_FILENAME,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ─────────────────────── VERSION HISTORY ───────────────────────

def show_versions():
    st.title("🕐 Version History")
    conn = get_db_connection()
    if st.session_state.role == "admin":
        versions = conn.execute(
            "SELECT v.*, u.username FROM excel_versions v JOIN users u ON v.user_id = u.id ORDER BY v.created_at DESC"
        ).fetchall()
    else:
        versions = conn.execute(
            "SELECT v.*, u.username FROM excel_versions v JOIN users u ON v.user_id = u.id WHERE v.user_id = ? ORDER BY v.created_at DESC",
            (st.session_state.user_id,),
        ).fetchall()
    conn.close()

    if not versions:
        st.info("No versions uploaded yet. Go to **Upload Data** to add your first file.")
        return

    for v in versions:
        v = dict(v)
        label = f"v{v['version_number']} — {v['filename']}  ({v['created_at'][:10]})  {'✅ Active' if v['is_active'] else ''}"
        with st.expander(label):
            col1, col2, col3 = st.columns(3)
            col1.write(f"**Uploaded by:** {v['username']}")
            col2.write(f"**Records:** {v['total_wheel_records']}")
            col3.write(f"**Status:** {'Active' if v['is_active'] else 'Inactive'}")

            version_path = get_version_storage_path(v["id"])
            btn1, btn2 = st.columns(2)

            if not v["is_active"] and os.path.exists(version_path):
                if btn1.button("🔄 Activate this version", key=f"activate_{v['id']}"):
                    shutil.copy2(version_path, DATA_PATH)
                    conn = get_db_connection()
                    conn.execute("UPDATE excel_versions SET is_active = 0 WHERE user_id = ?", (v["user_id"],))
                    conn.execute("UPDATE excel_versions SET is_active = 1 WHERE id = ?", (v["id"],))
                    conn.commit()
                    conn.close()
                    st.cache_data.clear()
                    st.success(f"Version {v['version_number']} activated!")
                    st.rerun()

            if os.path.exists(version_path):
                with open(version_path, "rb") as f:
                    btn2.download_button(
                        "⬇️ Download",
                        data=f.read(),
                        file_name=v["filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{v['id']}",
                    )


# ─────────────────────── USER MANAGEMENT ───────────────────────

def show_user_management():
    st.title("👥 User Management")
    if st.session_state.role != "admin":
        st.error("Admin access required.")
        return

    conn = get_db_connection()
    users = conn.execute(
        "SELECT id, username, role, is_active, created_at FROM users ORDER BY created_at DESC"
    ).fetchall()
    conn.close()

    st.subheader("All Users")
    users_df = pd.DataFrame([dict(u) for u in users])
    st.dataframe(users_df, use_container_width=True)

    st.divider()
    st.subheader("➕ Add New User")
    with st.form("add_user_form"):
        new_username = st.text_input("Username")
        new_password = st.text_input("Password", type="password")
        new_role = st.selectbox("Role", ["user", "admin"])
        if st.form_submit_button("Create User", type="primary"):
            if not new_username or not new_password:
                st.error("Username and password are required")
            else:
                try:
                    conn = get_db_connection()
                    conn.execute(
                        "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                        (new_username, hash_password(new_password), new_role),
                    )
                    conn.commit()
                    conn.close()
                    st.success(f"User '{new_username}' created successfully!")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Username already exists")

    st.divider()
    st.subheader("🗑️ Delete User")
    conn = get_db_connection()
    other_users = conn.execute(
        "SELECT id, username FROM users WHERE id != ?", (st.session_state.user_id,)
    ).fetchall()
    conn.close()

    if other_users:
        user_options = {f"{u['username']} (id={u['id']})": u["id"] for u in other_users}
        selected = st.selectbox("Select user to delete", list(user_options.keys()))
        if st.button("Delete Selected User", type="secondary"):
            uid = user_options[selected]
            conn = get_db_connection()
            conn.execute("DELETE FROM users WHERE id = ?", (uid,))
            conn.commit()
            conn.close()
            st.success("User deleted.")
            st.rerun()
    else:
        st.info("No other users to delete.")


# ─────────────────────── MAIN ───────────────────────

def main():
    init_db()

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    if not st.session_state.logged_in:
        show_login()
        return

    # Load data
    df_wheels, df_hardness, df_rail, using_real_data = load_data()

    # Sidebar
    with st.sidebar:
        st.markdown("### 🏗️ YZ Bay Analytics")
        st.markdown(f"👤 **{st.session_state.username}** `{st.session_state.role}`")
        st.divider()

        pages = [
            "📊 Dashboard",
            "🔬 Hardness Analysis",
            "⚙️ Wheel Failure Analysis",
            "📈 Failure Distribution",
            "🛤️ Rail Replacement Log",
            "🔮 Predictions",
            "📤 Upload Data",
            "🕐 Version History",
        ]
        if st.session_state.role == "admin":
            pages.append("👥 User Management")

        page = st.radio("Navigate", pages, label_visibility="collapsed")
        st.divider()

        if st.button("🚪 Logout", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    # Render selected page
    if page == "📊 Dashboard":
        show_dashboard(df_wheels, df_hardness, df_rail, using_real_data)
    elif page == "🔬 Hardness Analysis":
        show_hardness(df_hardness)
    elif page == "⚙️ Wheel Failure Analysis":
        show_wheel_failure(df_wheels)
    elif page == "📈 Failure Distribution":
        show_failure_distribution(df_wheels)
    elif page == "🛤️ Rail Replacement Log":
        show_rail_replacement(df_rail)
    elif page == "🔮 Predictions":
        show_predictions(df_wheels, df_hardness)
    elif page == "📤 Upload Data":
        show_upload()
    elif page == "🕐 Version History":
        show_versions()
    elif page == "👥 User Management":
        show_user_management()


if __name__ == "__main__":
    main()
